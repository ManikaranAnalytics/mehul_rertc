"""
Live formula-driven RE-RTC Dispatch Workbook
Edit RTC Commitment / WTG Count / Solar AC MW in Config → everything recalculates.
"""
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ── palette ───────────────────────────────────────────────────────────────────
DARK="0D1426"; HDR="1E2A45"; WIND="00B4D8"; SOLAR="F59E0B"
PSPD="8B5CF6"; PSPC="EC4899"; NET="10B981"; WARN="EF4444"; CURT="334155"

def _f(h): return PatternFill("solid", fgColor=h)
def _fn(b=False,c="E2E8F0",s=10,i=False): return Font(bold=b,color=c,size=s,name="Calibri",italic=i)
def _bd():
    t=Side(style="thin",color="1E2A45")
    return Border(left=t,right=t,top=t,bottom=t)
def _al(h="right",w=False): return Alignment(horizontal=h,vertical="center",wrap_text=w)
def _s(c,bg=DARK,fg="E2E8F0",b=False,s=10,a="right",i=False,f=None):
    c.fill=_f(bg); c.font=_fn(b,fg,s,i); c.border=_bd()
    c.alignment=_al(a,a=="center")
    if f: c.number_format=f

# Config sheet named cell refs (cross-sheet formulas)
RTC  = "Config!$B$6"
WTG  = "Config!$B$8"
SOL  = "Config!$B$11"
LOSS = "Config!$B$18"    # round-trip loss %
COMP = "Config!$B$19"    # min compliance ratio
CS   = "Config!$B$20"    # curtailment start block
CE   = "Config!$B$21"    # curtailment end block
MDSP = "Config!$B$22"    # min dispatch MW (CERC 6 MW rule)
MAXS = "Config!$B$23"    # max SoC MWh
ISOC = "Config!$B$24"    # initial SoC (EOD SoC from previous day)
DS   = "'Dispatch Schedule'"
RAW  = "'Raw Data'"


def build_excel(forecast_df, block_results, summary, rtc_range,
                rtc_commitment, wtg_count, solar_ac_mw, date_str,
                initial_soc_mwh=0.0,
                curtailment_enabled=True,
                curtailment_start_block=37, curtailment_end_block=64,
                curtailment_label="",
                psp_curtailment_label="None",
                roundtrip_loss_pct=20.0, min_compliance_ratio=0.50,
                max_soc_mwh=360.0, max_charge_mw=60.0, max_discharge_mw=50.0,
                min_dispatch_mw=6.0, transmission_loss_pct=0.0,
                discharge_target="rtc_commitment") -> bytes:
    wb = Workbook()
    ws_cfg  = wb.active;  ws_cfg.title = "Config"
    ws_raw  = wb.create_sheet("Raw Data")
    ws_disp = wb.create_sheet("Dispatch Schedule")
    ws_sum  = wb.create_sheet("Summary")
    _cfg(ws_cfg, date_str, wtg_count, solar_ac_mw, rtc_commitment, rtc_range,
         curtailment_enabled, curtailment_start_block, curtailment_end_block,
         curtailment_label, psp_curtailment_label,
         roundtrip_loss_pct, min_compliance_ratio, initial_soc_mwh,
         max_soc_mwh, max_charge_mw, max_discharge_mw, min_dispatch_mw,
         transmission_loss_pct, discharge_target)
    _raw(ws_raw, forecast_df, wtg_count, solar_ac_mw)
    _disp(ws_disp, block_results, rtc_commitment)
    _summ(ws_sum, summary, block_results, rtc_commitment, max_soc_mwh)
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf.read()


def _cfg(ws, date_str, wtg_count, solar_ac_mw, rtc, rtc_range,
         curtailment_enabled, curtailment_start_block, curtailment_end_block,
         curtailment_label, psp_curtailment_label,
         roundtrip_loss_pct, min_compliance_ratio, initial_soc_mwh=0.0,
         max_soc_mwh=360.0, max_charge_mw=60.0, max_discharge_mw=50.0,
         min_dispatch_mw=6.0, transmission_loss_pct=0.0,
         discharge_target="rtc_commitment"):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 44

    def title(span, v, fg="A5B4FC", s=13):
        ws.merge_cells(span); c=ws[span.split(":")[0]]
        c.value=v; _s(c,bg=DARK,fg=fg,b=True,s=s,a="center")

    def sec(r, v):
        ws.merge_cells(f"A{r}:C{r}"); c=ws.cell(r,1,v)
        _s(c,bg="1A2744",fg="818CF8",b=True,s=10,a="left")
        ws.row_dimensions[r].height=18

    def row(r, label, val, note="", edit=False, fg="F8FAFC", fmt=None):
        a=ws.cell(r,1,label); _s(a,bg=HDR,fg="94A3B8",b=True,s=10,a="left")
        b=ws.cell(r,2,val)
        _s(b,bg="0A1020" if edit else DARK,fg=fg,b=True,s=12 if edit else 10,a="right",f=fmt)
        if edit:
            g=Side(style="medium",color="10B981")
            b.border=Border(left=g,right=g,top=g,bottom=g)
        if note:
            nc=ws.cell(r,3,note); _s(nc,bg=DARK,fg="64748B",i=True,s=9,a="left")
        ws.row_dimensions[r].height=20

    title("A1:C1","RE-RTC Dispatch Optimizer — Live Configuration",s=14)
    ws.row_dimensions[1].height=30
    title("A2:C2","Aditya Birla Renewables  |  Hindalco Mahan 100 MW RTC Captive PPA",fg="64748B",s=10)
    ws.row_dimensions[2].height=16

    sec(4,"  ▸  SIMULATION INPUTS  (edit green-bordered cells — all sheets recalculate)")
    row(5,  "Simulation Date",           date_str)
    row(6,  "RTC Commitment (MW)  ✎",    rtc,
        "← EDIT THIS — flat daily commitment target",edit=True,fg="34D399",fmt="0.0")
    row(7,  "Min Compliance Floor (MW)", f"={RTC}*{COMP}",
        f"Auto: {int(min_compliance_ratio*100)}% of RTC — regulatory minimum")
    row(8,  "Wind Turbines (WTGs)  ✎",   wtg_count,
        "← EDIT THIS — active WTG count (1–59)",edit=True,fg="00D2FF",fmt="0")
    row(9,  "WTG Unit Capacity (MW)",    3.15,"Suzlon S144 3.0/3.15 MW (fixed)")
    row(10, "Total Wind Capacity (MW)",  f"={WTG}*3.15","Auto from WTG count")
    row(11, "Solar AC Capacity (MW)  ✎", solar_ac_mw,
        "← EDIT THIS — AC-side net capacity (5–175 MW)",edit=True,fg="F59E0B",fmt="0.0")

    sec(13,"  ▸  PSP STORAGE & DISPATCH CONSTANTS")
    row(14, "PSP Location",              "Orvakallu PSP, Andhra Pradesh")
    row(15, "Max Storage (MWh)",         max_soc_mwh,"PSP storage ceiling",fmt="0.0")
    row(16, "Max Charge Rate (MW)",      max_charge_mw, "Max draw from grid",fmt="0.0")
    row(17, "Max Discharge Rate (MW)",   max_discharge_mw, "Max injection to grid",fmt="0.0")
    row(18, "Round-Trip Loss (%)",       roundtrip_loss_pct,
        "Total energy loss charging + discharging",fmt="0.0")
    row(19, "Min Compliance Ratio",      min_compliance_ratio,
        f"{int(min_compliance_ratio*100)}% of RTC is the regulatory floor",fmt="0.00")
    curt_display = curtailment_label or (
        f"B{curtailment_start_block}–{curtailment_end_block}" if curtailment_enabled else "DISABLED"
    )
    row(20, "Wind+Solar Curtailment",    curt_display,
        "Blocks where wind/solar generation is zeroed")
    row(21, "PSP Discharge Curtailment", psp_curtailment_label,
        "Per-block PSP discharge caps (0 = external supply)")
    row(22, "Min Dispatch MW (CERC)",    min_dispatch_mw,
        "CERC rule: PSP must dispatch 0 or ≥ this value; sub-threshold bumped up",fmt="0.0")
    row(23, "Transmission Loss (%)",     transmission_loss_pct, fmt="0.0")
    row(24, "Discharge Target Mode",     discharge_target)

    sec(26,"  ▸  CARRY-FORWARD FROM PREVIOUS DAY")
    row(27, "Initial SoC (MWh)",        initial_soc_mwh,
        "EOD SoC from previous day — sets Block 1 SoC Start in Dispatch Schedule",
        fmt="0.0")

    if rtc_range and "min_rtc_mw" in rtc_range:
        sec(29,"  ▸  MANIKARAN'S SUGGESTION  (dispatch-validated commitment analysis)")
        st=rtc_range.get("generation_stats",{})
        row(30,"Min Safe Commit (MW)",       rtc_range["min_rtc_mw"],
            f"{int(min_compliance_ratio*100)}% of P10 non-curtail gen",fg=WARN,fmt="0.00")
        row(31,"★ Recommended Commit (MW)",  rtc_range["recommended_rtc_mw"],
            "Max RTC → zero shortfall across all 96 blocks (incl. 6MW CERC bump)",fg="34D399",fmt="0.00")
        row(32,"Max Aggressive (MW)",        rtc_range["max_rtc_mw"],
            "P90 non-curtail gen (PSP backup for low blocks)",fg="818CF8",fmt="0.00")
        row(33,"Non-Curtail Gen P10 (MW)",   st.get("p10_mw",""),fmt="0.00")
        row(34,"Non-Curtail Gen Mean (MW)",  st.get("mean_mw",""),fmt="0.00")
        row(35,"Non-Curtail Gen P90 (MW)",   st.get("p90_mw",""),fmt="0.00")
        row(36,"Curtailment Loss (MWh/day)", rtc_range.get("curtailment_period_gen_lost_mwh",""),fmt="0.00")

    sec(38,"  ▸  HOW TO USE THIS WORKBOOK")
    ws.merge_cells("A39:C42")
    c=ws.cell(39,1)
    c.value=("1. Config sheet records the inputs used for this export.\n"
             "2. Dispatch Schedule shows the optimized 96-block result (matches the app).\n"
             "3. Summary aggregates daily KPIs from the dispatch run.\n"
             "4. Raw Data holds meteorological / uploaded generation source data.")
    _s(c,bg="0A1830",fg="94A3B8",i=True,s=10,a="left")
    c.alignment=Alignment(horizontal="left",vertical="top",wrap_text=True)
    ws.row_dimensions[39].height=70


def _raw(ws, forecast_df, wtg_count, solar_ac_mw):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"
    for col,w in {"A":7,"B":9,"C":20,"D":20,"E":14,"F":11}.items():
        ws.column_dimensions[col].width=w

    ws.merge_cells("A1:F1"); c=ws["A1"]
    c.value="Raw Data — Wind & Solar Source Data  (do NOT edit — auto-generated by backend)"
    _s(c,bg=DARK,fg="A5B4FC",b=True,s=11,a="center")
    ws.row_dimensions[1].height=22

    for col,lbl,clr in [("A","Block","94A3B8"),("B","Time","94A3B8"),
                         ("C","Per-WTG Power (kW)",WIND),
                         ("D","Solar Fraction\n(at any Solar AC MW)",SOLAR),
                         ("E","Wind Speed\nProjected (m/s)","7DD3FC"),
                         ("F","Curtailed?","64748B")]:
        c=ws[f"{col}2"]; c.value=lbl
        _s(c,bg=HDR,fg=clr,b=True,s=9,a="center")
    ws.row_dimensions[2].height=28

    for i,rd in enumerate(forecast_df.to_dict("records")):
        r=i+3
        is_c=bool(rd.get("curtail_flag",False))
        bg=CURT if is_c else ("111827" if i%2==0 else DARK)

        def rc(col,val,fmt=None,clr="CBD5E1",a="right"):
            c=ws[f"{col}{r}"]; c.value=val
            _s(c,bg=bg,fg=clr,s=9,a=a,f=fmt)

        wind_raw   = float(rd.get("wind_mw_raw",0))
        solar_raw  = float(rd.get("solar_mw_raw",0))
        per_wtg_kw = (wind_raw/wtg_count*1000.0) if wtg_count>0 else 0.0
        sol_frac   = (solar_raw/solar_ac_mw)      if solar_ac_mw>0 else 0.0

        rc("A",int(rd["block"]),              clr="94A3B8",a="center")
        rc("B",str(rd.get("time",""))[:5],    clr="94A3B8",a="center")
        rc("C",round(per_wtg_kw,4),           fmt="0.0000",clr=WIND)
        rc("D",round(sol_frac,6),             fmt="0.000000",clr=SOLAR)
        rc("E",round(float(rd.get("wind_speed",0)),2),fmt="0.00",clr="7DD3FC")
        rc("F","YES" if is_c else "NO",
           clr="F59E0B" if is_c else NET,a="center")


def _disp(ws, block_results, rtc_commitment):
    """Write actual optimized dispatch values (matches on-screen schedule)."""
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    COLS=[("A","Block",6,"94A3B8"),("B","Time",9,"94A3B8"),
          ("C","Curtailed?",11,"64748B"),
          ("D","Wind MW",12,WIND),
          ("E","Solar MW",13,SOLAR),
          ("F","Combined\nGen MW",12,"22D3EE"),
          ("G","Min Floor\nMW",11,WARN),
          ("H","SoC Start\nMWh",11,"6366F1"),
          ("I","PSP\nDischarge MW",13,PSPD),
          ("J","PSP\nCharge MW",12,PSPC),
          ("K","SoC End\nMWh",11,"6366F1"),
          ("L","Net Schedule\nMW",13,NET),
          ("M","RTC Target\nMW",11,WARN),
          ("N","RTM Surplus\nMW",12,"6B7280"),
          ("O","Compliant?",10,"334155")]

    ws.merge_cells("A1:O1"); c=ws["A1"]
    c.value="Dispatch Schedule — 96 Time Blocks  |  Optimized result (matches app)"
    _s(c,bg=DARK,fg="A5B4FC",b=True,s=12,a="center")
    ws.row_dimensions[1].height=24

    for col,lbl,w,clr in COLS:
        ws.column_dimensions[col].width=w
        c=ws[f"{col}2"]; c.value=lbl
        _s(c,bg=HDR,fg=clr,b=True,s=9,a="center")
    ws.row_dimensions[2].height=32

    for i, b in enumerate(block_results):
        r=i+3
        is_c = bool(b.get("curtail_flag", False))
        bg=CURT if is_c else ("111827" if i%2==0 else DARK)

        def fc(col,val,fmt=None,fg="CBD5E1",a="right",bold=False):
            c=ws[f"{col}{r}"]; c.value=val
            _s(c,bg=bg,fg=fg,s=9,a=a,f=fmt,b=bold)

        fc("A", int(b["block"]), a="center", fg="94A3B8")
        fc("B", str(b.get("time", ""))[:5], a="center", fg="94A3B8")
        fc("C", "YES" if is_c else "NO", fg="F59E0B" if is_c else NET, a="center", bold=True)
        fc("D", float(b.get("wind_mw", 0)), fmt="0.00", fg=WIND)
        fc("E", float(b.get("solar_mw", 0)), fmt="0.00", fg=SOLAR)
        fc("F", float(b.get("generation_mw", 0)), fmt="0.00", fg="22D3EE", bold=True)
        fc("G", float(b.get("min_schedule", 0)), fmt="0.00", fg=WARN)
        fc("H", float(b.get("soc_start", 0)), fmt="0.0", fg="818CF8")
        fc("I", float(b.get("psp_discharge", 0)), fmt="0.00", fg=PSPD)
        fc("J", float(b.get("psp_charge", 0)), fmt="0.00", fg=PSPC)
        fc("K", float(b.get("soc_end", 0)), fmt="0.0", fg="818CF8")
        fc("L", float(b.get("net_schedule", 0)), fmt="0.00", fg=NET, bold=True)
        fc("M", rtc_commitment, fmt="0.00", fg=WARN)
        fc("N", float(b.get("rtm_surplus", 0)), fmt="0.00", fg="6B7280")
        compliant = bool(b.get("compliant", False))
        fc("O", "✓ YES" if compliant else "✗ NO", fg=NET, a="center", bold=True)

    # Totals row 99
    ws.merge_cells("A99:B99"); c=ws.cell(99,1,"TOTALS / AVERAGES")
    _s(c,bg=HDR,fg="A5B4FC",b=True,s=9,a="center")
    totals = {
        "D": round(sum(float(b.get("wind_mw", 0)) for b in block_results) / 96, 2),
        "E": round(sum(float(b.get("solar_mw", 0)) for b in block_results) / 96, 2),
        "F": round(sum(float(b.get("generation_mw", 0)) for b in block_results) / 96, 2),
        "I": round(sum(float(b.get("psp_discharge", 0)) for b in block_results) * 0.25, 2),
        "J": round(sum(float(b.get("psp_charge", 0)) for b in block_results) * 0.25, 2),
        "K": round(float(block_results[-1].get("soc_end", 0)) if block_results else 0, 1),
        "L": round(sum(float(b.get("net_schedule", 0)) for b in block_results) / 96, 2),
        "N": round(sum(float(b.get("rtm_surplus", 0)) for b in block_results) * 0.25, 2),
        "O": f"{sum(1 for b in block_results if b.get('compliant'))} / 96",
    }
    for col, val in totals.items():
        c=ws[f"{col}99"]; c.value=val
        fmt = "0.00" if col != "K" and col != "O" else ("0.0" if col == "K" else None)
        _s(c,bg=HDR,fg="F8FAFC",b=True,s=9,a="right",f=fmt)
    ws.row_dimensions[99].height=18


def _summ(ws, summary, block_results, rtc_commitment, max_soc_mwh=360.0):
    """Write daily KPIs from the optimization summary dict."""
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width=36
    ws.column_dimensions["B"].width=24
    ws.column_dimensions["C"].width=36

    ws.merge_cells("A1:C1"); c=ws["A1"]
    c.value="Daily Dispatch Summary — from optimized schedule (matches app)"
    _s(c,bg=DARK,fg="A5B4FC",b=True,s=13,a="center")
    ws.row_dimensions[1].height=28

    def sec(r,v):
        ws.merge_cells(f"A{r}:C{r}"); c=ws.cell(r,1,v)
        _s(c,bg="1A2744",fg="818CF8",b=True,s=10,a="left")
        ws.row_dimensions[r].height=18

    def row(r,label,val,note="",fg="F8FAFC",fmt=None):
        a=ws.cell(r,1,label); _s(a,bg=HDR,fg="94A3B8",b=True,s=10,a="left")
        b=ws.cell(r,2,val); _s(b,bg=DARK,fg=fg,b=True,s=11,a="right",f=fmt)
        if note:
            nc=ws.cell(r,3,note); _s(nc,bg=DARK,fg="64748B",i=True,s=9,a="left")
        ws.row_dimensions[r].height=20

    s = summary or {}
    total_wind = round(sum(float(b.get("wind_mw", 0)) for b in block_results) * 0.25, 2)
    total_solar = round(sum(float(b.get("solar_mw", 0)) for b in block_results) * 0.25, 2)
    total_gen = round(sum(float(b.get("generation_mw", 0)) for b in block_results) * 0.25, 2)
    avg_gen = round(sum(float(b.get("generation_mw", 0)) for b in block_results) / 96, 2) if block_results else 0

    sec(3,"  ▸  CONFIGURATION")
    row(4, "RTC Commitment (MW)",         rtc_commitment, fmt="0.00", fg="34D399")
    row(5, "Min Compliance Floor (MW)",   s.get("min_schedule_mw", rtc_commitment * 0.5), fmt="0.00")
    row(6, "Min Compliance Ratio",        s.get("min_compliance_ratio", 0.5), fmt="0.0%")
    row(7, "Discharge Target Mode",       s.get("discharge_target", "rtc_commitment"))

    sec(11,"  ▸  GENERATION TOTALS")
    row(12,"Total Wind Generation (MWh)", total_wind, fmt="0.00", fg=WIND)
    row(13,"Total Solar Generation (MWh)",total_solar, fmt="0.00", fg=SOLAR)
    row(14,"Total Combined Gen (MWh)",    total_gen, fmt="0.00", fg="22D3EE")
    row(15,"Avg Gen per Block (MW)",      avg_gen, fmt="0.00")

    sec(17,"  ▸  PSP STORAGE DISPATCH")
    row(18,"Total PSP Discharged (MWh)",  s.get("total_discharged_mwh", 0), fmt="0.00", fg=PSPD)
    row(19,"Total PSP Charged (MWh)",     s.get("total_charged_mwh", 0), fmt="0.00", fg=PSPC)
    row(20,"PSP Usable Energy (MWh)",     s.get("psp_usable_charged_mwh", 0),
        "Actual recoverable energy after round-trip losses", fmt="0.00", fg=PSPC)
    row(21,"PSP Cycles Used",             s.get("cycles_used", 0), fmt="0.00")
    row(22,"End-of-Day SoC (MWh)",        s.get("end_soc_mwh", 0), fmt="0.0", fg="818CF8")
    end_soc = float(s.get("end_soc_mwh", 0) or 0)
    row(23,"End-of-Day SoC (%)",          end_soc / max_soc_mwh if max_soc_mwh else 0, fmt="0.0%", fg="818CF8")

    sec(25,"  ▸  COMPLIANCE & DELIVERY")
    compliant = int(s.get("compliant_blocks", 0))
    row(26,"Compliant Blocks",            f"{compliant} / 96", fg=NET)
    row(27,"Compliance Rate (%)",         compliant / 96 if block_results else 0, fmt="0.0%", fg=NET)
    row(28,"Total Net Schedule (MWh)",  s.get("total_net_delivered_mwh", 0), fmt="0.00", fg=NET)
    row(29,"Total RTM Surplus (MWh)",   s.get("total_rtm_surplus_mwh", 0), fmt="0.00", fg="6B7280",
        note="Exportable generation above RTC target")
    row(30,"Shortfall Energy (MWh)",      s.get("shortfall_energy_mwh", 0), fmt="0.00", fg=WARN)
    row(31,"Fully Compliant Day?",
        "✓  YES — 100% blocks met" if s.get("fully_compliant") else "✗  NO — shortfall blocks exist",
        fg=NET)

    sec(33,"  ▸  CARRY-FORWARD (PSP SoC Roll)")
    row(34,"Initial SoC (MWh)",           s.get("initial_soc_mwh", 0),
        "EOD SoC from previous day", fmt="0.0", fg="A78BFA")
    row(35,"Carry Forward Discharged (MWh)", s.get("carry_forward_discharged_mwh", 0),
        "Carry energy consumed today", fmt="0.0", fg="A78BFA")
    row(36,"End-of-Day SoC → next day carry", s.get("end_soc_mwh", 0),
        "Pass this value as Initial SoC for next day's simulation", fmt="0.0", fg="A78BFA")

    if s.get("charge_window_expired_mwh", 0):
        sec(38,"  ▸  24h CHARGE WINDOW")
        row(39,"Window Charged (MWh)",      s.get("charge_window_charged_mwh", 0), fmt="0.00")
        row(40,"Window Discharged (MWh)",   s.get("charge_window_discharged_mwh", 0), fmt="0.00")
        row(41,"Window Expired (MWh)",      s.get("charge_window_expired_mwh", 0), fmt="0.00", fg=WARN)
        row(42,"Window Outstanding (MWh)", s.get("charge_window_outstanding_mwh", 0), fmt="0.00")
